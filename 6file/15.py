import openpyxl

wb = openpyxl.load_workbook("trekking1.xlsx")
sheet = wb.active
listProduc = [];
for i in range(2,sheet.max_row+1):
    listProduc.append([sheet.cell(row = i,column = 1).value,sheet.cell(row = i,column = 2).value,])
fout = open('out15.txt','w',encoding = 'utf8')
for var in sorted(listProduc,key = lambda x: (-x[1],x[0])):
    print(var[0],file = fout)
fout.close()    


pass
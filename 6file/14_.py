import openpyxl
wb = openpyxl.load_workbook('salaries14.xlsx')
sheet = wb.active
for i in range(2,sheet.max_row+1):
    for j in range(2,sheet.max_column+1):
        d= sheet.cell(row = i, column = j)
        print(d.value,end = "\t")
    print()



pass
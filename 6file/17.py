import openpyxl

wb = openpyxl.load_workbook("trekking3.xlsx")
sheet1 = wb[wb.sheetnames[0]]
sheet2 = wb[wb.sheetnames[1]]
dictProduct = dict()
for row in sheet1:
    dictProduct[row[0].value] = [
        row[1].value or 0,
        row[2].value or 0,
        row[3].value or 0,
        row[4].value or 0,
    ]

valueOnDay = {}
for i in range(2, sheet2.max_row + 1):
    day = sheet2.cell(row=i, column=1).value
    if day not in valueOnDay:
        valueOnDay[day] = [0,0,0,0]
    product = sheet2.cell(row=i, column=2).value
    for j in range(4):
        valueOnDay[day][j] += (dictProduct[product][j] * sheet2.cell(row=i, column=3).value/100)

for day in valueOnDay:
    rec =list(map(int,valueOnDay[day]))
    print(*rec)

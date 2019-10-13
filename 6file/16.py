import openpyxl

wb = openpyxl.load_workbook("trekking2.xlsx")
# sheets = wb.sheetnames
sheet1 = wb[wb.sheetnames[0]]
sheet2 = wb[wb.sheetnames[1]]
# print(sheet1['A2'].value)
# print(sheet2['A2'].value)
dictProduct = dict()
for row in sheet1:
    # print(row)
    dictProduct[row[0].value] = [
        row[1].value or 0,
        row[2].value or 0,
        row[3].value or 0,
        row[4].value or 0,
    ]

value = [0, 0, 0, 0]

for i in range(2, sheet2.max_row + 1):
    product = sheet2.cell(row=i, column=1).value
    for j in range(4):
        value[j] += (dictProduct[product][j] * sheet2.cell(row=i, column=2).value/100)
    pass

for val in map(int,value):
    print(val,end = " ")
print()

pass

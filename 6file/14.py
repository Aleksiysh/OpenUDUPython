import openpyxl

wb = openpyxl.load_workbook("salaries14.xlsx")
sheet = wb.active
listAll = list()
for i in range(2, sheet.max_row + 1):
    listAll.append(list())
    for j in range(2, sheet.max_column + 1):
        d = sheet.cell(row=i, column=j)
        listAll[i - 2].append(d.value)

maxMedianValueIndex = 0
maxMedianValue = 0
listSumm = [0 for value in range(sheet.max_row - 2)]
maxAvgIndex = 0
for i in range(len(listAll)):
    # tmp = sorted(i)
    # listSumm.append(0)
    MedianValue = sorted(listAll[i])[len(listAll[i]) // 2]
    if MedianValue > maxMedianValue:
        maxMedianValue = MedianValue
        maxMedianValueIndex = i
    for j in range(len(listAll[i])):
        listSumm[j] += listAll[i][j]

for j in range(len(listSumm)):
    if listSumm[maxAvgIndex] < listSumm[j]:
        maxAvgIndex = j

print(sheet.cell(row=maxMedianValueIndex + 2, column=1).value, end=" ")
print(sheet.cell(row=1, column=maxAvgIndex + 2).value)
